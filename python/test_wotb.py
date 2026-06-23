#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
回归测试: 用 Data/ 下的真实回放断言解析/去重/汇总/导出不退化。

    python test_wotb.py

无需 pytest, 直接运行; 全部通过则退出码 0, 否则非 0。
"""
import os
import sys
import glob
import shutil
import tempfile

sys.stdout.reconfigure(encoding="utf-8") if hasattr(sys.stdout, "reconfigure") else None

import wotb_extractor as W

DATA_DIR = os.path.normpath(
    os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "common", "data"))
REPLAYS = sorted(glob.glob(os.path.join(DATA_DIR, "*.wotbreplay")))

passed = 0
failed = 0


def check(cond, msg):
    global passed, failed
    if cond:
        passed += 1
        print(f"  ✓ {msg}")
    else:
        failed += 1
        print(f"  ✗ 失败: {msg}")


def test_parse(path):
    print(f"[解析] {os.path.basename(path)}")
    battle, players = W.parse_replay(path)
    check(len(players) == 14, f"玩家数=14 (实际 {len(players)})")
    check(all(p["account_id"] for p in players), "每位玩家都有账号ID")
    check(battle.get("arena_id") is not None, "有 arenaUniqueId")

    # 字段关系: 发射 >= 命中 >= 击穿
    ok_shots = all(p["n_shots"] >= p["n_hits_dealt"] >= p["n_penetrations_dealt"] for p in players)
    check(ok_shots, "发射 ≥ 命中 ≥ 击穿 (全员)")
    # 被命中 = 被跳弹(#13) + 被击穿(#15)
    ok_recv = all(
        p["n_hits_received"] == W.f_uint(p["_raw"], 13) + p["n_penetrations_received"]
        for p in players)
    check(ok_recv, "被命中 = 被跳弹 + 被击穿 (全员)")

    # 击杀总数 == 阵亡人数
    total_kills = sum(p["kills"] for p in players)
    deaths = sum(1 for p in players if not p["survived"])
    check(total_kills == deaths, f"击杀总数({total_kills}) == 阵亡数({deaths})")
    # 强不变量(全歼或时间到都成立): 每队击杀数 == 敌队阵亡数
    deaths_by_team = {
        t: sum(1 for p in players if p["team"] == t and not p["survived"])
        for t in (1, 2)
    }
    kills_by_team = {
        t: sum(p["kills"] for p in players if p["team"] == t)
        for t in (1, 2)
    }
    check(kills_by_team[1] == deaths_by_team[2] and kills_by_team[2] == deaths_by_team[1],
          "每队击杀数 == 敌队阵亡数")
    return battle, players


def test_dedup_and_aggregate():
    print("[去重 + 汇总]")
    tmp = tempfile.mkdtemp(prefix="wotb_test_")
    try:
        # 复制全部回放, 再额外复制一份制造"重复上传"
        for p in REPLAYS:
            shutil.copy(p, tmp)
        dup = os.path.join(tmp, "duplicate_upload.wotbreplay")
        shutil.copy(REPLAYS[0], dup)

        files = sorted(glob.glob(os.path.join(tmp, "*.wotbreplay")))
        battles, dups, fails = W.collect_battles(files)
        check(len(battles) == len(REPLAYS), f"唯一战斗={len(REPLAYS)} (实际 {len(battles)})")
        check(len(dups) == 1, f"识别出 1 个重复 (实际 {len(dups)})")
        check(len(fails) == 0, "无解析失败")

        tp = W.load_tankopedia()
        agg = W.aggregate_players(battles, tp)
        # 每位选手场次 = 其出现的唯一战斗数, 上限为战斗总数
        check(all(1 <= a["battles"] <= len(battles) for a in agg.values()),
              "每位选手场次在 [1, 战斗数] 内")
        # 胜场 <= 场次
        check(all(a["wins"] <= a["battles"] for a in agg.values()), "胜场 ≤ 场次")

        out = os.path.join(tmp, "agg.xlsx")
        W.export_aggregate_xlsx(battles, tp, out, dups)
        from openpyxl import load_workbook
        wb = load_workbook(out)
        check(wb.sheetnames == ["汇总", "明细", "战斗列表"], "汇总工作簿含 3 个表")
        check(wb["明细"].max_row - 1 == 14 * len(battles), "明细行数 = 14 × 战斗数")
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


def test_single_export(path):
    print("[单场导出]")
    tp = W.load_tankopedia()
    battle, players = W.parse_replay(path)
    tmp = tempfile.mkdtemp(prefix="wotb_test_")
    try:
        out = os.path.join(tmp, "single.xlsx")
        W.export_xlsx(battle, players, tp, out)
        from openpyxl import load_workbook
        wb = load_workbook(out)
        check(wb.sheetnames == ["战斗信息", "玩家数据", "原始字段"], "单场工作簿含 3 个表")
        hdr = [c.value for c in wb["玩家数据"][1]]
        expect = [c.title for c in W.PLAYER_COLUMNS]
        check(hdr == expect, "玩家数据列与 PLAYER_COLUMNS 一致")
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


def test_tankopedia():
    print("[车辆库]")
    tp = W.load_tankopedia()
    check(len(tp) > 600, f"车辆库非空 ({len(tp)} 辆)")
    check(W.tank_info(tp, 4481)["name"] == "Kranvagn", "4481 = Kranvagn")
    # 轻坦车种不为空(回归: 之前 class=0 被省略导致空白)
    empties = [k for k, v in tp.items() if not v.get("class")]
    check(not empties, f"所有车都有车种 (空: {len(empties)})")


def main():
    if not REPLAYS:
        print("Data/ 下没有 .wotbreplay, 无法测试。")
        return 1
    test_tankopedia()
    for path in REPLAYS:
        test_parse(path)
    test_single_export(REPLAYS[0])
    test_dedup_and_aggregate()
    print(f"\n结果: {passed} 通过, {failed} 失败")
    return 0 if failed == 0 else 1


if __name__ == "__main__":
    sys.exit(main())
