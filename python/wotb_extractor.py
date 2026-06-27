#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
WoT Blitz 回放数据提取工具 (wotb_extractor)
============================================

从 .wotbreplay 文件中提取一场战斗里全部 14 名玩家的各项数据，并导出为 xlsx。

用法:
    wotb_extractor.exe <回放文件.wotbreplay> [-o 输出.xlsx]
    wotb_extractor.exe <包含回放的文件夹>            # 批量处理
    直接把 .wotbreplay 文件拖到 exe 上也可以。

数据来源:
    .wotbreplay 本身是一个 zip 包, 其中:
      - meta.json          : 战斗基本信息(地图/版本/时长/录像者车辆等)
      - battle_results.dat  : pickle( arenaId, protobuf(BattleResults) )
                              protobuf 内含每位玩家详细战绩。

protobuf 字段含义参考社区项目 eigenein/wotbreplay-parser, 并以本回放交叉验证。
"""

import sys
import os
import json
import zipfile
import pickle
import struct
import argparse
import datetime
import collections

# Windows 控制台默认 cp1252, 重配为 UTF-8 以正确输出中文
for _stream in (sys.stdout, sys.stderr):
    try:
        _stream.reconfigure(encoding="utf-8")
    except (AttributeError, ValueError):
        pass

# ---------------------------------------------------------------------------
# 资源路径 (兼容 PyInstaller --onefile 打包)
# ---------------------------------------------------------------------------
def resource_path(rel):
    """定位打包资源。
    - 冻结(PyInstaller)时: 资源被 --add-data 放在 _MEIPASS 根。
    - 源码运行时: 共享资源(如 tankopedia.json)在仓库的 common/ 目录,
      即 python/ 的同级 ../common/; 兼容旧的同目录放置。
    """
    meipass = getattr(sys, "_MEIPASS", None)
    if meipass:
        return os.path.join(meipass, rel)
    here = os.path.dirname(os.path.abspath(__file__))
    candidates = [
        os.path.join(here, rel),                       # 同目录(兼容旧布局)
        os.path.join(here, "..", "common", rel),       # 新布局: common/
    ]
    for path in candidates:
        if os.path.exists(path):
            return path
    return candidates[0]


# ---------------------------------------------------------------------------
# 通用 protobuf 解码器 (无需 .proto / protoc)
# 返回 dict: field_number -> [value, ...]  (重复字段保留多值)
# 叶子 length-delimited 字段保留为 bytes, 由调用方决定按字符串还是嵌套消息解释。
# ---------------------------------------------------------------------------
def _read_varint(buf, i):
    shift = 0
    result = 0
    while True:
        b = buf[i]
        i += 1
        result |= (b & 0x7F) << shift
        if not (b & 0x80):
            break
        shift += 7
    return result, i


def decode_protobuf(buf):
    """把一段 protobuf 字节解码为 {field: [values]}。
    length-delimited 字段: 先尝试当作嵌套消息解析(若能完整消费), 否则保留 bytes。"""
    fields = {}
    i = 0
    n = len(buf)
    while i < n:
        try:
            tag, i = _read_varint(buf, i)
        except IndexError:
            break
        field = tag >> 3
        wt = tag & 7
        if field == 0:
            break
        try:
            if wt == 0:  # varint
                val, i = _read_varint(buf, i)
            elif wt == 1:  # 64-bit
                val = struct.unpack("<Q", buf[i:i + 8])[0]
                i += 8
            elif wt == 5:  # 32-bit
                val = struct.unpack("<I", buf[i:i + 4])[0]
                i += 4
            elif wt == 2:  # length-delimited
                ln, i = _read_varint(buf, i)
                val = buf[i:i + ln]
                i += ln
            else:
                # 3/4 = group(已废弃), 6/7 = 非法 -> 停止
                break
        except (IndexError, struct.error):
            break
        fields.setdefault(field, []).append(val)
    return fields


def as_str(raw):
    """length-delimited 字段当作字符串解释 (优先 UTF-8)。"""
    if isinstance(raw, str):
        return raw
    if not isinstance(raw, (bytes, bytearray)):
        return raw
    for enc in ("utf-8", "latin1"):
        try:
            return raw.decode(enc)
        except UnicodeDecodeError:
            continue
    return raw.hex()


def f1(fields, num, default=None):
    """取字段的第一个值; 不存在返回 default。"""
    v = fields.get(num)
    if not v:
        return default
    return v[0]


def f_uint(fields, num, default=0):
    v = f1(fields, num, None)
    return default if v is None else v


# ---------------------------------------------------------------------------
# tankopedia (tank_id -> 名称/等级/类型/国家), 数据来自 blitzkit, 2026-06
# 字段已本地化为中文, 见 tankopedia.json 的 meta。
# ---------------------------------------------------------------------------
def load_tankopedia():
    try:
        with open(resource_path("tankopedia.json"), encoding="utf-8") as fp:
            data = json.load(fp)
        return data.get("data", data)
    except Exception:
        return {}


def tank_info(tankopedia, tank_id):
    t = tankopedia.get(str(tank_id))
    if not t:
        return {"name": f"#{tank_id}", "tier": "", "type": "", "nation": ""}
    return {
        "name": t.get("name") or f"#{tank_id}",
        "tier": t.get("tier", ""),
        "type": t.get("class", ""),
        "nation": t.get("nation", ""),
    }


# ---------------------------------------------------------------------------
# 解析 .wotbreplay
#
# protobuf 字段号集中在此 (语义只存一处, 便于核对/维护):
#   battle_results.dat = pickle( (arenaId, protobuf BattleResults) )
#   BattleResults: #3=获胜队, #201=玩家名册(repeated), #301=玩家战绩(repeated)
#   Player(#201):        #1=account_id, #2=PlayerInfo
#   PlayerInfo:          见 ROSTER_FIELDS
#   PlayerResults(#301): #1=result_id, #2=PlayerResultsInfo
#   PlayerResultsInfo:   见 RESULT_UINT_FIELDS / ASSIST_FIELDS / FIELD_*
# ---------------------------------------------------------------------------
TEAM_NAME = {1: "队伍1", 2: "队伍2"}
UINT_NEG1 = 0xFFFFFFFFFFFFFFFF  # 18446744073709551615 = -1, 表示存活(无人击杀)

# 名册 PlayerInfo (#201 -> #2) 字段
ROSTER_FIELDS = {"nickname": 1, "platoon_id": 2, "team": 3, "clan": 5, "rank": 9}
ROSTER_STR_KEYS = ("nickname", "clan")  # 这些按字符串(UTF-8)解释, 其余按数值

# 战绩 PlayerResultsInfo (#301 -> #2) 的简单 uint 字段: 输出键 -> protobuf 字段号
RESULT_UINT_FIELDS = {
    "account_id": 101, "team": 102, "tank_id": 103,
    "n_shots": 4, "n_hits_dealt": 5, "n_penetrations_dealt": 7,
    "damage_dealt": 8, "damage_received": 11,
    "n_hits_received": 12, "n_penetrations_received": 15,
    "n_enemies_damaged": 17, "kills": 18,
    "damage_blocked": 117,
    # 含义存疑, 仍解析备用 (不在主表展示):
    "xp": 23, "credits": 106,
}
ASSIST_FIELDS = (9, 10)   # 辅助伤害 = damage_assisted_1 + _2
FIELD_SURVIVED = 105      # == UINT_NEG1 表示存活
FIELD_MM_RATING = 107     # float 位模式, 仅排位赛有意义


def parse_replay(path):
    with zipfile.ZipFile(path) as zf:
        names = zf.namelist()
        meta = {}
        if "meta.json" in names:
            meta = json.loads(zf.read("meta.json").decode("utf-8"))
        if "battle_results.dat" not in names:
            raise ValueError("回放中没有 battle_results.dat (可能是不完整或加密的回放)")
        dat = zf.read("battle_results.dat")

    arena_id, pb_bytes = pickle.loads(dat, encoding="bytes")
    root = decode_protobuf(pb_bytes)

    # ---- 名册: #201 -> Player(#1 account_id, #2 info) ----
    roster = {}
    for praw in root.get(201, []):
        p = decode_protobuf(praw)
        info = decode_protobuf(f1(p, 2, b""))
        roster[f1(p, 1)] = {
            key: (as_str(f1(info, num, b"")) if key in ROSTER_STR_KEYS else f1(info, num))
            for key, num in ROSTER_FIELDS.items()
        }

    # ---- 详细战绩: #301 -> PlayerResults(#1 result_id, #2 info) ----
    players = []
    for rraw in root.get(301, []):
        r = decode_protobuf(rraw)
        info = decode_protobuf(f1(r, 2, b""))
        rec = {key: f_uint(info, num) for key, num in RESULT_UINT_FIELDS.items()}
        rec["result_id"] = f1(r, 1)
        rec["damage_assisted"] = sum(f_uint(info, num) for num in ASSIST_FIELDS)
        # 逐击杀目标明细尚未解析时, 潜在伤害先保守等于实际伤害。
        rec["potential_damage"] = rec.get("damage_dealt", 0)
        rec["potential_damage_supplement"] = 0
        rec["potential_damage_detail"] = "未解析"
        rec["_raw"] = dict(info)
        # 存活判断: 仅幸存者带有 #105 == -1(全F) 标记; 阵亡者无此字段
        rec["survived"] = (f1(info, FIELD_SURVIVED) == UINT_NEG1)
        # 排位等级分: #107 是 float 的位模式
        mm = f1(info, FIELD_MM_RATING)
        if mm is not None:
            try:
                rec["mm_rating"] = struct.unpack("<f", struct.pack("<I", mm & 0xFFFFFFFF))[0]
                rec["display_rating"] = round(rec["mm_rating"] * 10.0 + 3000.0)
            except struct.error:
                rec["mm_rating"] = None
        players.append(rec)

    # 合并名册信息
    for rec in players:
        info = roster.get(rec["account_id"], {})
        rec["nickname"] = info.get("nickname") or str(rec["account_id"])
        rec["clan"] = info.get("clan") or ""
        rec["platoon_id"] = info.get("platoon_id")
        rec["rank"] = info.get("rank")

    battle = {
        "arena_id": arena_id,
        "winner_team": f1(root, 3),
        "mode_map_id": f1(root, 1),
        "version": meta.get("version", ""),
        "map_name": meta.get("mapName", ""),
        "map_id": meta.get("mapId", ""),
        "duration_s": meta.get("battleDuration"),
        "start_time": meta.get("battleStartTime"),
        "recorder": meta.get("playerName", ""),
        "recorder_vehicle": meta.get("playerVehicleName", ""),
        "n_players": len(players),
    }
    return battle, players


# ---------------------------------------------------------------------------
# 显示派生字段 + 列定义 (Excel 与 GUI 共用, 避免重复)
# ---------------------------------------------------------------------------
Col = collections.namedtuple("Col", "title key xlsx px num")
# title=表头  key=数据键  xlsx=Excel列宽  px=GUI列宽(像素)  num=是否数值(右对齐/数值排序)

# 身份/车辆列
IDENTITY_COLUMNS = [
    Col("玩家", "nickname", 20, 130, False),
    Col("战队", "clan", 10, 70, False),
    Col("车辆", "tank_name", 20, 120, False),
    Col("等级", "tank_tier", 6, 45, True),
    Col("坦克类型", "tank_type", 9, 60, False),
    Col("国家", "tank_nation", 8, 55, False),
]
# 战斗数据列 (玩家数据表 / 明细表 共用)
STAT_COLUMNS = [
    Col("评分", "rating", 6, 55, True),
    Col("存活", "survived_label", 6, 45, False),
    Col("击杀", "kills", 6, 45, True),
    Col("伤害", "damage_dealt", 8, 65, True),
    Col("潜在伤害", "potential_damage", 9, 70, True),
    Col("补增伤害", "potential_damage_supplement", 9, 70, True),
    Col("潜在明细", "potential_damage_detail", 9, 65, False),
    Col("协助伤害", "damage_assisted", 9, 65, True),
    Col("损失血量", "damage_received", 9, 65, True),
    Col("格挡", "damage_blocked", 9, 65, True),
    Col("发射", "n_shots", 6, 45, True),
    Col("命中", "n_hits_dealt", 6, 45, True),
    Col("击穿", "n_penetrations_dealt", 6, 45, True),
    Col("被命中", "n_hits_received", 7, 50, True),
    Col("被击穿", "n_penetrations_received", 7, 50, True),
    Col("击伤", "n_enemies_damaged", 9, 55, True),
]
# 尾部附加列
TAIL_COLUMNS = [
    Col("排", "platoon_label", 6, 40, False),
    Col("车辆ID", "tank_id", 9, 65, True),
    Col("账号ID", "account_id", 12, 95, True),
]
# 单场「玩家数据」表的完整列
PLAYER_COLUMNS = IDENTITY_COLUMNS + STAT_COLUMNS + TAIL_COLUMNS
# 左对齐(文本)列, 其余居中
LEFT_ALIGN_KEYS = {"nickname", "clan", "tank_name", "date", "map_name"}


def enrich_display(player, tankopedia):
    """给玩家记录补上展示用派生字段(车名/车种/队伍/存活标签)。原地修改并返回。"""
    ti = tank_info(tankopedia, player.get("tank_id"))
    player["tank_name"] = ti["name"]
    player["tank_tier"] = ti["tier"]
    player["tank_type"] = ti["type"]
    player["tank_nation"] = ti["nation"]
    player["team_label"] = TEAM_NAME.get(player.get("team"), player.get("team"))
    player["survived_label"] = "存活" if player.get("survived") else "阵亡"
    return player


def make_platoon_labeler():
    """返回一个把 platoon_id 映射成 A/B/C… 的函数(每次调用独立计数)。"""
    letters = {}

    def label(pid):
        if not pid:
            return ""
        if pid not in letters:
            letters[pid] = chr(ord("A") + len(letters))
        return letters[pid]

    return label


def sort_players(players):
    """统一排序: 先队伍, 同队按伤害降序。"""
    return sorted(players, key=lambda r: (r.get("team", 0), -r.get("damage_dealt", 0)))


# ---------------------------------------------------------------------------
# 导出 xlsx
# ---------------------------------------------------------------------------
def _excel_styles():
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    thin = Side(style="thin", color="BFBFBF")
    return {
        "hdr_font": Font(bold=True, color="FFFFFF"),
        "hdr_fill": PatternFill("solid", fgColor="2F5597"),
        "team_fill": {1: PatternFill("solid", fgColor="DDEBF7"),
                      2: PatternFill("solid", fgColor="FCE4D6")},
        "center": Alignment(horizontal="center", vertical="center"),
        "border": Border(left=thin, right=thin, top=thin, bottom=thin),
    }


def _display_width(text):
    """按中文(全角)算 2、其余算 1 的显示宽度。"""
    return sum(2 if ord(ch) > 0x2E7F else 1 for ch in str(text))


def _write_header(ws, columns, st):
    """写表头(支持 Col 具名元组或 (title, key, width) 三元组)。
    列宽自动取 max(设定宽度, 表头所需宽度), 避免被自动筛选下拉箭头截断。"""
    from openpyxl.utils import get_column_letter
    for c, col in enumerate(columns, start=1):
        title = col.title if isinstance(col, Col) else col[0]
        width = (col.xlsx if isinstance(col, Col) else col[2]) or 0
        # 表头中文宽度 + 自动筛选下拉箭头(约 4 字宽)余量
        width = max(width, _display_width(title) + 4)
        cell = ws.cell(row=1, column=c, value=title)
        cell.font = st["hdr_font"]
        cell.fill = st["hdr_fill"]
        cell.alignment = st["center"]
        cell.border = st["border"]
        ws.column_dimensions[get_column_letter(c)].width = width


def export_xlsx(battle, players, tankopedia, out_path):
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    compute_ratings([(battle, players, None)], tankopedia)   # 基准=该场内
    wb = Workbook()
    st = _excel_styles()

    # ---------- Sheet 1: 战斗信息 ----------
    ws = wb.active
    ws.title = "战斗信息"
    start = battle.get("start_time")
    try:
        start_str = datetime.datetime.fromtimestamp(int(start)).strftime("%Y-%m-%d %H:%M:%S")
    except (TypeError, ValueError):
        start_str = str(start)
    dur = battle.get("duration_s")
    dur_str = f"{int(dur // 60)}分{int(dur % 60)}秒" if isinstance(dur, (int, float)) else str(dur)
    win = battle.get("winner_team")
    info_rows = [
        ("游戏版本", battle.get("version")),
        ("地图", battle.get("map_name")),
        ("开始时间", start_str),
        ("战斗时长", dur_str),
        ("获胜队伍", TEAM_NAME.get(win, "平局/未知" if win in (0, None) else win)),
        ("录像者", battle.get("recorder")),
        ("录像者车辆", battle.get("recorder_vehicle")),
        ("玩家数", battle.get("n_players")),
        ("竞技场ID", battle.get("arena_id")),
    ]
    ws["A1"] = "战斗信息"
    ws["A1"].font = Font(bold=True, size=14)
    for idx, (k, v) in enumerate(info_rows, start=3):
        ws.cell(row=idx, column=1, value=k).font = Font(bold=True)
        ws.cell(row=idx, column=2, value=v)
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 40

    # ---------- Sheet 2: 玩家数据 ----------
    ws = wb.create_sheet("玩家数据")
    players_sorted = sort_players(players)
    platoon_label = make_platoon_labeler()
    for r in players_sorted:
        enrich_display(r, tankopedia)
        r["platoon_label"] = platoon_label(r.get("platoon_id"))

    _write_header(ws, PLAYER_COLUMNS, st)
    for ridx, r in enumerate(players_sorted, start=2):
        fill = st["team_fill"].get(r["team"])
        for c, col in enumerate(PLAYER_COLUMNS, start=1):
            cell = ws.cell(row=ridx, column=c, value=r.get(col.key, ""))
            if fill:
                cell.fill = fill
            cell.border = st["border"]
            if col.key not in LEFT_ALIGN_KEYS:
                cell.alignment = st["center"]
    ws.freeze_panes = "B2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(PLAYER_COLUMNS))}{len(players_sorted) + 1}"

    # ---------- Sheet 3: 原始字段 (完整透明) ----------
    ws = wb.create_sheet("原始字段")
    all_fields = set()
    for r in players:
        all_fields |= set(r["_raw"].keys())
    field_cols = sorted(all_fields)
    ws.cell(row=1, column=1, value="玩家").font = Font(bold=True)
    ws.cell(row=1, column=2, value="账号ID").font = Font(bold=True)
    for c, fn in enumerate(field_cols, start=3):
        cell = ws.cell(row=1, column=c, value=f"#{fn}")
        cell.font = Font(bold=True)
    for ridx, r in enumerate(sorted(players, key=lambda x: (x["team"], -x["damage_dealt"])), start=2):
        ws.cell(row=ridx, column=1, value=r.get("nickname"))
        ws.cell(row=ridx, column=2, value=r.get("account_id"))
        for c, fn in enumerate(field_cols, start=3):
            vals = r["_raw"].get(fn)
            if not vals:
                continue
            disp = ", ".join(v.hex() if isinstance(v, (bytes, bytearray)) else str(v) for v in vals)
            ws.cell(row=ridx, column=c, value=disp)

    wb.save(out_path)
    return out_path


# ---------------------------------------------------------------------------
# 多回放: 按 arenaUniqueId 去重 + 跨场次按账号汇总
# ---------------------------------------------------------------------------
def collect_battles(paths, on_log=None):
    """解析多个回放, 按 arenaUniqueId 去重(同一场战斗/重复上传只算一次)。

    返回 (battles, duplicates, failures):
      battles    : [(battle, players, path), ...]  去重后的唯一战斗
      duplicates : [(path, arena_id), ...]          被跳过的重复
      failures   : [(path, 错误信息), ...]
    """
    def log(msg):
        if on_log:
            on_log(msg)

    seen = {}            # arena_id -> path
    battles = []
    duplicates = []
    failures = []
    for path in paths:
        try:
            battle, players = parse_replay(path)
        except Exception as e:
            failures.append((path, str(e)))
            log(f"[失败] {os.path.basename(path)}: {e}")
            continue
        aid = battle.get("arena_id")
        if aid in seen:
            duplicates.append((path, aid))
            log(f"[跳过-重复] {os.path.basename(path)} (与 {os.path.basename(seen[aid])} 同一场)")
            continue
        seen[aid] = path
        battles.append((battle, players, path))
        log(f"[读取] {os.path.basename(path)}  地图:{battle.get('map_name')}  玩家:{len(players)}")
    return battles, duplicates, failures


def _safe_div(a, b):
    return (a / b) if b else 0.0


# ---------------------------------------------------------------------------
# 评分 (自包含, 按车型基准归一化; 与 Java Rating.java 一致)
# ---------------------------------------------------------------------------
RATING_W_ASSIST = 0.6
RATING_W_BLOCK = 0.35
RATING_KILL_VALUE = 200
RATING_WIN_BONUS = 0.05
RATING_MIN_SAMPLES = 5
RATING_SCALE = 1000


def effective_contribution(r):
    """有效贡献(伤害当量)。"""
    return (r["damage_dealt"]
            + RATING_W_ASSIST * r["damage_assisted"]
            + RATING_W_BLOCK * r["damage_blocked"]
            + RATING_KILL_VALUE * r["kills"])


def compute_ratings(battles, tankopedia):
    """对一批战斗的所有玩家写入 r['rating']; 基准按车型从这批数据求得。"""
    by_class = {}            # cls -> [sumEC, count]
    all_sum = 0.0
    all_n = 0
    for _battle, players, _ in battles:
        for r in players:
            ec = effective_contribution(r)
            cls = tank_info(tankopedia, r["tank_id"])["type"] or "其他"
            acc = by_class.setdefault(cls, [0.0, 0])
            acc[0] += ec
            acc[1] += 1
            all_sum += ec
            all_n += 1
    if all_n == 0:
        return
    overall = all_sum / all_n
    for battle, players, _ in battles:
        winner = battle.get("winner_team")
        for r in players:
            cls = tank_info(tankopedia, r["tank_id"])["type"] or "其他"
            acc = by_class.get(cls)
            baseline = acc[0] / acc[1] if (acc and acc[1] >= RATING_MIN_SAMPLES) else overall
            if baseline <= 0:
                baseline = overall if overall > 0 else 1
            ratio = effective_contribution(r) / baseline
            win = bool(winner) and r["team"] == winner
            r["rating"] = round(RATING_SCALE * ratio * (1 + (RATING_WIN_BONUS if win else 0)))


def aggregate_players(battles, tankopedia):
    """跨场次按账号ID累计每位选手的数据。battles 来自 collect_battles。"""
    agg = {}
    for battle, players, _ in battles:
        winner = battle.get("winner_team")
        start = battle.get("start_time")
        try:
            start = int(start)
        except (TypeError, ValueError):
            start = 0
        for r in players:
            acc = r["account_id"]
            a = agg.get(acc)
            if a is None:
                a = agg[acc] = {
                    "account_id": acc, "nickname": "", "clan": "", "_last": -1,
                    "battles": 0, "wins": 0, "survived": 0,
                    "kills": 0, "damage": 0, "assisted": 0, "received": 0, "blocked": 0,
                    "shots": 0, "hits": 0, "pens": 0,
                    "hits_received": 0, "pens_received": 0, "enemies_damaged": 0,
                    "rating_sum": 0, "tanks": {},
                }
            # 用最近一场的昵称/战队
            if start >= a["_last"]:
                a["_last"] = start
                a["nickname"] = r.get("nickname") or str(acc)
                a["clan"] = r.get("clan") or ""
            a["battles"] += 1
            if winner and r["team"] == winner:
                a["wins"] += 1
            if r.get("survived"):
                a["survived"] += 1
            a["kills"] += r["kills"]
            a["damage"] += r["damage_dealt"]
            a["assisted"] += r["damage_assisted"]
            a["received"] += r["damage_received"]
            a["blocked"] += r["damage_blocked"]
            a["shots"] += r["n_shots"]
            a["hits"] += r["n_hits_dealt"]
            a["pens"] += r["n_penetrations_dealt"]
            a["hits_received"] += r["n_hits_received"]
            a["pens_received"] += r["n_penetrations_received"]
            a["enemies_damaged"] += r["n_enemies_damaged"]
            a["rating_sum"] += r.get("rating") or 0
            tname = tank_info(tankopedia, r["tank_id"])["name"]
            a["tanks"][tname] = a["tanks"].get(tname, 0) + 1
    return agg


def export_aggregate_xlsx(battles, tankopedia, out_path, duplicates=None):
    """把多场战斗合并导出为一个汇总工作簿: 汇总 / 明细 / 战斗列表。"""
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    st = _excel_styles()
    border = st["border"]
    center = st["center"]

    def write_header(ws, cols):
        _write_header(ws, cols, st)

    # ---------- Sheet 1: 汇总 (每位选手跨场累计) ----------
    compute_ratings(battles, tankopedia)   # 基准=这批战斗
    agg = aggregate_players(battles, tankopedia)
    ws = wb.active
    ws.title = "汇总"
    agg_cols = [
        ("玩家", "nickname", 18),
        ("战队", "clan", 10),
        ("场次", "battles", 6),
        ("胜场", "wins", 6),
        ("胜率%", "win_rate", 8),
        ("存活率%", "survival_rate", 9),
        ("场均评分", "rating_avg", 8),
        ("总击杀", "kills", 7),
        ("场均击杀", "kills_avg", 7),
        ("总伤害", "damage", 9),
        ("场均伤害", "damage_avg", 9),
        ("总协助伤害", "assisted", 9),
        ("场均协助伤害", "assisted_avg", 9),
        ("场均损失血量", "received_avg", 8),
        ("场均格挡", "blocked_avg", 8),
        ("命中率%", "hit_rate", 8),
        ("击穿率%", "pen_rate", 8),
        ("场均击伤", "enemies_damaged_avg", 9),
        ("用车", "tanks_str", 30),
        ("账号ID", "account_id", 12),
    ]
    write_header(ws, agg_cols)
    rows = []
    for a in agg.values():
        n = a["battles"]
        tanks_str = ", ".join(f"{name}×{cnt}" if cnt > 1 else name
                              for name, cnt in sorted(a["tanks"].items(), key=lambda x: -x[1]))
        rows.append({
            **a,
            "win_rate": round(_safe_div(a["wins"], n) * 100, 1),
            "survival_rate": round(_safe_div(a["survived"], n) * 100, 1),
            "rating_avg": round(_safe_div(a["rating_sum"], n)),
            "kills_avg": round(_safe_div(a["kills"], n), 2),
            "damage_avg": round(_safe_div(a["damage"], n), 1),
            "assisted_avg": round(_safe_div(a["assisted"], n), 1),
            "received_avg": round(_safe_div(a["received"], n), 1),
            "blocked_avg": round(_safe_div(a["blocked"], n), 1),
            "hit_rate": round(_safe_div(a["hits"], a["shots"]) * 100, 1),
            "pen_rate": round(_safe_div(a["pens"], a["shots"]) * 100, 1),
            "enemies_damaged_avg": round(_safe_div(a["enemies_damaged"], n), 2),
            "tanks_str": tanks_str,
        })
    # 默认按 均伤害 降序排名
    rows.sort(key=lambda r: -r["damage_avg"])
    for ridx, r in enumerate(rows, start=2):
        for c, (_t, key, _w) in enumerate(agg_cols, start=1):
            cell = ws.cell(row=ridx, column=c, value=r.get(key, ""))
            cell.border = border
            if c >= 3:
                cell.alignment = center
    ws.freeze_panes = "B2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(agg_cols))}{len(rows) + 1}"

    # ---------- Sheet 2: 明细 (每场 × 每人) ----------
    ws = wb.create_sheet("明细")
    # 复用战斗数据列(STAT_COLUMNS), 前面加场次信息, 末尾加账号
    detail_cols = [
        Col("日期", "date", 17, 0, False),
        Col("地图", "map_name", 12, 0, False),
        Col("玩家", "nickname", 16, 0, False),
        Col("战队", "clan", 9, 0, False),
        Col("车辆", "tank_name", 16, 0, False),
        Col("胜负", "result", 6, 0, False),
    ] + STAT_COLUMNS + [Col("账号ID", "account_id", 12, 0, True)]
    write_header(ws, detail_cols)
    ridx = 2
    for battle, players, _ in battles:
        try:
            date = datetime.datetime.fromtimestamp(int(battle.get("start_time"))).strftime("%Y-%m-%d %H:%M")
        except (TypeError, ValueError):
            date = ""
        winner = battle.get("winner_team")
        for r in sort_players(players):
            row = enrich_display(dict(r), tankopedia)
            row["date"] = date
            row["map_name"] = battle.get("map_name")
            row["result"] = "胜" if (winner and r["team"] == winner) else ("负" if winner else "平")
            for c, col in enumerate(detail_cols, start=1):
                cell = ws.cell(row=ridx, column=c, value=row.get(col.key, ""))
                cell.border = border
                if col.key not in LEFT_ALIGN_KEYS:
                    cell.alignment = center
            ridx += 1
    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(detail_cols))}{ridx - 1}"

    # ---------- Sheet 3: 战斗列表 ----------
    ws = wb.create_sheet("战斗列表")
    bl_cols = [
        ("序号", None, 6), ("日期", None, 17), ("地图", None, 12),
        ("时长", None, 9), ("获胜队", None, 8), ("玩家数", None, 7),
        ("arenaUniqueId", None, 22), ("文件名", None, 40),
    ]
    write_header(ws, bl_cols)
    ridx = 2
    for i, (battle, players, path) in enumerate(battles, start=1):
        try:
            date = datetime.datetime.fromtimestamp(int(battle.get("start_time"))).strftime("%Y-%m-%d %H:%M:%S")
        except (TypeError, ValueError):
            date = ""
        dur = battle.get("duration_s")
        dur_s = f"{int(dur // 60)}分{int(dur % 60)}秒" if isinstance(dur, (int, float)) else dur
        win = battle.get("winner_team")
        vals = [i, date, battle.get("map_name"), dur_s,
                TEAM_NAME.get(win, "平局/未知"), len(players),
                str(battle.get("arena_id")), os.path.basename(path)]
        for c, v in enumerate(vals, start=1):
            ws.cell(row=ridx, column=c, value=v).border = border
        ridx += 1
    # 重复文件附在底部
    if duplicates:
        ridx += 1
        ws.cell(row=ridx, column=1, value="已跳过的重复上传:").font = Font(bold=True, color="C00000")
        ridx += 1
        for path, aid in duplicates:
            ws.cell(row=ridx, column=2, value=os.path.basename(path))
            ws.cell(row=ridx, column=7, value=str(aid))
            ridx += 1

    wb.save(out_path)
    return out_path, agg


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------
def process_one(path, out=None):
    battle, players = parse_replay(path)
    tankopedia = load_tankopedia()
    if out is None:
        out = os.path.splitext(path)[0] + ".xlsx"
    export_xlsx(battle, players, tankopedia, out)
    return out, battle, players


def main(argv=None):
    parser = argparse.ArgumentParser(
        description="WoT Blitz 回放数据提取工具 — 导出 14 名玩家战绩到 xlsx")
    parser.add_argument("input", nargs="+", help="一个或多个 .wotbreplay 文件, 或包含回放的文件夹")
    parser.add_argument("-o", "--output", help="输出 xlsx 路径")
    parser.add_argument("--each", action="store_true",
                        help="多个回放时, 每个各自导出一个 xlsx (默认是合并为一个汇总工作簿)")
    args = parser.parse_args(argv)

    targets = []
    for inp in args.input:
        if os.path.isdir(inp):
            for fn in sorted(os.listdir(inp)):
                if fn.lower().endswith(".wotbreplay"):
                    targets.append(os.path.join(inp, fn))
        elif inp.lower().endswith(".wotbreplay"):
            targets.append(inp)

    if not targets:
        print("未找到任何 .wotbreplay 文件。")
        return 1

    # 单个文件, 或显式 --each: 每个回放各自导出
    if len(targets) == 1 or args.each:
        ok = 0
        for path in targets:
            try:
                out = args.output if (len(targets) == 1 and args.output) else None
                out_path, battle, players = process_one(path, out)
                print(f"[OK] {os.path.basename(path)}")
                print(f"     地图: {battle['map_name']}  时长: {battle['duration_s']}  玩家: {len(players)}")
                print(f"     -> {out_path}")
                ok += 1
            except Exception as e:
                print(f"[失败] {os.path.basename(path)}: {e}")
        print(f"\n完成: {ok}/{len(targets)} 个回放已导出。")
        return 0 if ok else 2

    # 多个文件: 去重 + 合并为一个汇总工作簿
    battles, duplicates, failures = collect_battles(targets, on_log=print)
    if not battles:
        print("没有可用的回放。")
        return 2
    tankopedia = load_tankopedia()
    out = args.output or os.path.join(
        os.path.dirname(os.path.abspath(targets[0])), "联赛汇总.xlsx")
    out_path, agg = export_aggregate_xlsx(battles, tankopedia, out, duplicates)
    print(f"\n唯一战斗: {len(battles)}  |  跳过重复: {len(duplicates)}  |  "
          f"失败: {len(failures)}  |  选手: {len(agg)}")
    print(f"汇总已导出 -> {out_path}")
    return 0


if __name__ == "__main__":
    code = main()
    # 双击/拖拽运行时, 暂停以便查看结果
    if sys.stdout and sys.stdout.isatty() and os.environ.get("WOTB_NO_PAUSE") != "1":
        try:
            input("\n按回车键退出...")
        except (EOFError, KeyboardInterrupt):
            pass
    sys.exit(code)
